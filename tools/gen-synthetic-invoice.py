#!/usr/bin/env python3
"""
Generates a synthetic Azure Cost analysis export (.xlsx) that joins
cleanly to the Contoso DMC Azure-mode scan fixture.

The output mirrors the Azure portal's "Cost analysis → Download → CSV/Excel"
shape exactly:

  Summary sheet:
    rows of ('', '<label>', '<value>', None) — Name, Type, ID, View,
    Start date, End date, Granularity, Group by, Actual cost
  Data sheet:
    columns: ResourceId, ResourceType, ResourceLocation, ResourceGroupName,
             SubscriptionName, ServiceName, Meter, Tags, CostUSD, Cost,
             Currency

Each VM in the DMC scan gets matching invoice rows (compute + license +
disks + bandwidth) keyed by an Azure-format ResourceId so a downstream
join on `/subscriptions/.../virtualMachines/<vm_name>` resolves directly
to its DMC `machine_id`. Non-VM rows cover the resource categories from
`resource_scan_summary.results`.

Run:
    python3 tools/gen-synthetic-invoice.py [--scan PATH] [--out PATH]
"""
from __future__ import annotations

import argparse
import json
import random
import re
from datetime import date, datetime, timedelta
from pathlib import Path

import openpyxl
from openpyxl.utils import get_column_letter

SEED = "contoso-invoice-2026-q1"
rng = random.Random(SEED)

# ---------------------------------------------------------------------------
# Pricing model — approximate Azure PAYG list prices (USD).
# Scope: enough fidelity to drive JJ rules; not authoritative.
# ---------------------------------------------------------------------------
VM_HOURLY_USD = {
    "Standard_B2s":     0.0416,
    "Standard_B4ms":    0.166,
    "Standard_D4s_v5":  0.192,
    "Standard_D8s_v5":  0.384,
    "Standard_D16s_v5": 0.768,
    "Standard_D32s_v5": 1.536,
    "Standard_E8ds_v5": 0.504,
    "Standard_E16ds_v5": 1.008,
    "Standard_E32ds_v5": 2.016,
}
WINDOWS_LICENSE_PER_CORE_HOUR_USD = 0.046

DISK_PRICE_PER_GB_MONTH_USD = {
    "Standard_LRS":      0.04,
    "StandardSSD_LRS":   0.075,
    "Premium_LRS":       0.135,
}
HOURS_PER_MONTH = 730

REGION_DISPLAY = {"eastus": "EastUS", "westeurope": "WestEurope"}


def jitter(rate: float, lo: float = 0.92, hi: float = 1.08) -> float:
    return rate * rng.uniform(lo, hi)


def az_id(sub: str, rg: str, provider: str, name: str, *path) -> str:
    base = f"/subscriptions/{sub}/resourceGroups/{rg}/providers/{provider}/{name}"
    return base + ("/" + "/".join(path) if path else "")


def emit_vm_rows(vm: dict, sub: str, sub_name: str, currency: str, period_days: int):
    """Produce invoice rows for one DMC VM: compute + license + disks + bandwidth."""
    rows = []
    rg = vm["azure_resource_group"]
    name = vm["server_name"]
    region = vm["azure_region"]
    region_display = REGION_DISPLAY.get(region, region)
    size = vm["vm_size"]
    cores = vm["number_of_cores"]
    powered_on_hours = vm["cpu"]["powered_on_hours"] if vm.get("cpu") else 0
    # Scale 30-day-ish DMC `powered_on_hours` to the invoice period.
    if vm.get("metric_collection", {}).get("duration_days"):
        d = vm["metric_collection"]["duration_days"] or 1
        scaled_hours = powered_on_hours * (period_days / d)
    else:
        scaled_hours = powered_on_hours

    rid = f"/subscriptions/{sub}/resourceGroups/{rg}/providers/Microsoft.Compute/virtualMachines/{name}"
    tags = ""

    # 1. VM compute hours
    if scaled_hours > 0 and size in VM_HOURLY_USD:
        rate = jitter(VM_HOURLY_USD[size])
        cost = round(rate * scaled_hours, 6)
        rows.append([rid, "microsoft.compute/virtualmachines", region_display, rg, sub_name,
                     "Virtual Machines", f"{size} {region_display}", tags,
                     cost, cost, currency])

    # 2. Windows license uplift (visible separately on the invoice — Jeannie Rule 2)
    if vm.get("os_type", "").lower() == "windows" and scaled_hours > 0:
        cost = round(WINDOWS_LICENSE_PER_CORE_HOUR_USD * cores * scaled_hours, 6)
        rows.append([rid, "microsoft.compute/virtualmachines", region_display, rg, sub_name,
                     "Virtual Machines Licenses", f"Windows Server {cores} vCPU License",
                     tags, cost, cost, currency])

    # 3. Managed disks — per-disk row (exactly as Azure does per LUN)
    for d in vm.get("disk_data", {}).get("disks", {}).values():
        storage = d.get("storage_type") or "Standard_LRS"
        gb = d.get("capacity_gb") or 0
        rate_per_gb_month = DISK_PRICE_PER_GB_MONTH_USD.get(storage, 0.04)
        # Disks bill 24/7 regardless of VM power state.
        cost = round(jitter(rate_per_gb_month) * gb * (period_days / 30.0), 6)
        disk_rid = f"/subscriptions/{sub}/resourceGroups/{rg}/providers/Microsoft.Compute/disks/{d['disk_name']}"
        rows.append([disk_rid, "microsoft.compute/disks", region_display, rg, sub_name,
                     "Storage", f"{storage} Managed Disks {gb} GB",
                     tags, cost, cost, currency])

    # 4. Bandwidth (egress) — small line, only when VM was running.
    egress_gb = vm.get("network_data", {}).get("metrics", {}).get("aggregate", {}).get(
        "net.transmitted.total_gb", 0
    )
    if egress_gb and scaled_hours > 0:
        # Scale total_gb to invoice period as well.
        scaled_egress = egress_gb * (period_days / max(d_or_one(vm), 1))
        # First 100 GB free, then ~$0.087/GB.
        chargeable = max(0.0, scaled_egress - 100.0)
        cost = round(chargeable * 0.087, 6)
        if cost > 0.01:
            rows.append([rid, "microsoft.network/publicipaddresses", region_display, rg, sub_name,
                         "Bandwidth", "Inter-Region Data Transfer Out", tags,
                         cost, cost, currency])
    return rows


def d_or_one(vm: dict) -> int:
    return vm.get("metric_collection", {}).get("duration_days") or 1


# ---------------------------------------------------------------------------
# Non-VM resources — billed against the resource_scan_summary inventory.
# Rates are monthly; we scale to the period.
# ---------------------------------------------------------------------------
NON_VM_CATALOG = [
    # (resource_type_for_id, provider_path, service_name, meter, monthly_lo, monthly_hi)
    ("storage_accounts",        ("Microsoft.Storage", "storageAccounts"),
        "Storage", "General Purpose v2 Hot LRS", 5, 90),
    ("web_apps",                ("Microsoft.Web", "sites"),
        "Azure App Service", "P1v3 App Service", 110, 220),
    ("app_service_plans",       ("Microsoft.Web", "serverfarms"),
        "Azure App Service", "P1v3 App Service Plan", 140, 380),
    ("application_gateways",    ("Microsoft.Network", "applicationGateways"),
        "Application Gateway", "Standard_v2 Gateway hours", 220, 480),
    ("nat_gateways",            ("Microsoft.Network", "natGateways"),
        "Virtual Network", "NAT Gateway hours", 32, 95),
    ("azure_firewalls",         ("Microsoft.Network", "azureFirewalls"),
        "Azure Firewall", "Premium Firewall hours", 880, 1300),
    ("front_doors",             ("Microsoft.Cdn", "profiles"),
        "Content Delivery Network", "Premium Front Door", 190, 360),
    ("public_ip_addresses",     ("Microsoft.Network", "publicIPAddresses"),
        "Virtual Network", "Standard Static Public IP", 3, 5),
    ("load_balancers",          ("Microsoft.Network", "loadBalancers"),
        "Load Balancer", "Standard LB", 22, 60),
    ("key_vaults",              ("Microsoft.KeyVault", "vaults"),
        "Key Vault", "Standard Operations", 1, 12),
    ("sql_servers",             ("Microsoft.Sql", "servers"),
        "SQL Database", "GP_Gen5 8 vCore", 600, 1800),
    ("sql_managed_instances",   ("Microsoft.Sql", "managedInstances"),
        "SQL Managed Instance", "GP MI 8 vCore", 1900, 3600),
    ("postgresql_servers",      ("Microsoft.DBforPostgreSQL", "flexibleServers"),
        "Azure Database for PostgreSQL", "GP D4ds_v5", 220, 540),
    ("cosmosdb_accounts",       ("Microsoft.DocumentDB", "databaseAccounts"),
        "Azure Cosmos DB", "Provisioned Throughput RU/s", 320, 1200),
    ("recovery_vaults",         ("Microsoft.RecoveryServices", "vaults"),
        "Backup", "Azure VM Backup Storage", 14, 180),
    ("event_hub_namespaces",    ("Microsoft.EventHub", "namespaces"),
        "Event Hubs", "Standard Throughput Unit", 22, 120),
    ("servicebus_namespaces",   ("Microsoft.ServiceBus", "namespaces"),
        "Service Bus", "Premium Messaging Unit", 660, 700),
    ("redis_caches",            ("Microsoft.Cache", "Redis"),
        "Azure Cache for Redis", "Premium P1 6GB", 410, 440),
    ("container_registries",    ("Microsoft.ContainerRegistry", "registries"),
        "Container Registry", "Premium Registry", 35, 50),
    ("aks_clusters",            ("Microsoft.ContainerService", "managedClusters"),
        "Azure Kubernetes Service", "Free Tier control plane", 0, 5),
    ("vmss",                    ("Microsoft.Compute", "virtualMachineScaleSets"),
        "Virtual Machine Scale Sets", "Standard_D4s_v5 nodes", 240, 720),
    ("application_insights",    ("Microsoft.Insights", "components"),
        "Application Insights", "Data Ingestion", 6, 38),
    ("log_analytics_workspaces", ("Microsoft.OperationalInsights", "workspaces"),
        "Log Analytics workspace", "Pay-as-you-go data ingestion", 22, 110),
    ("private_dns_zones",       ("Microsoft.Network", "privateDnsZones"),
        "Virtual Network", "Private DNS Zone", 0.5, 1.5),
    ("dns_zones",               ("Microsoft.Network", "dnsZones"),
        "DNS", "Public DNS Zone", 0.5, 0.5),
    ("network_security_groups", ("Microsoft.Network", "networkSecurityGroups"),
        "Virtual Network", "NSG Rule Hours", 0, 0),  # billed at $0 — emitted for completeness
]


def emit_inventory_rows(scan, sub, sub_name, currency, period_days, primary_rg):
    rows = []
    counts = {k: v.get("count", 0) for k, v in
              scan["resource_scan_summary"]["results"].items() if isinstance(v, dict)}
    region_display = REGION_DISPLAY.get(scan["vm_results"][0]["azure_region"], "EastUS")
    for kind, (provider, sub_path), service, meter, lo, hi in NON_VM_CATALOG:
        n = counts.get(kind, 0)
        for i in range(n):
            monthly = jitter(rng.uniform(lo, hi), 0.85, 1.15) if hi > 0 else 0
            cost = round(monthly * (period_days / 30.0), 6)
            if cost <= 0:
                continue
            rg = pick_rg_for(kind, primary_rg, i)
            inst_name = f"{kind}-{i:03d}"
            rid = f"/subscriptions/{sub}/resourceGroups/{rg}/providers/{provider}/{sub_path}/{inst_name}"
            type_str = f"{provider.lower()}/{sub_path.lower()}"
            rows.append([rid, type_str, region_display, rg, sub_name, service, meter,
                         "", cost, cost, currency])
    return rows


# Distribute non-VM resources across the same RG namespace the VMs use.
RG_POOL = [
    "rg-contoso-prod-web-eus",
    "rg-contoso-prod-app-eus",
    "rg-contoso-prod-data-eus",
    "rg-contoso-uat-eus",
    "rg-contoso-dev-eus",
    "rg-contoso-shared-eus",
    "rg-contoso-prod-app-weu",
]


def pick_rg_for(kind: str, primary: str, i: int) -> str:
    return RG_POOL[(hash(kind) + i) % len(RG_POOL)]


# ---------------------------------------------------------------------------
# Sheet writers
# ---------------------------------------------------------------------------
def write_summary(ws, customer: str, sub_name: str, sub_id: str,
                  start: date, end: date, total: float):
    def row(label, value):
        ws.append(("", label, value, None))

    def blank():
        ws.append(("", None, None, None))

    blank()
    ws.append(("", "Scope", None, None))
    row("Name:", customer)
    row("Type:", "Subscription")
    row("ID:", f"/subscriptions/{sub_id}")
    blank()
    row("View:", "Custom view")
    row("Start date:", start.strftime("%a, %b %d, %Y"))
    row("End date:", end.strftime("%a, %b %d, %Y"))
    row("Granularity:", "None")
    row("Group by:", "ResourceId")
    blank()
    row("Actual cost:", round(total, 6))


HEADERS = [
    "ResourceId", "ResourceType", "ResourceLocation", "ResourceGroupName",
    "SubscriptionName", "ServiceName", "Meter", "Tags", "CostUSD", "Cost",
    "Currency",
]


def write_data(ws, rows):
    ws.append(HEADERS)
    for r in rows:
        ws.append(r)
    # Light formatting — widen columns for readability when opened.
    widths = [88, 38, 16, 30, 30, 26, 38, 12, 14, 14, 10]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


# ---------------------------------------------------------------------------
def main():
    ap = argparse.ArgumentParser()
    here = Path(__file__).resolve().parent.parent
    default_scan = here / "test-fixtures" / "dmc-azure-contoso"
    default_out = here / "test-fixtures" / "Contoso_Azure_Invoice_2026Q1.xlsx"
    ap.add_argument("--scan", type=Path, default=default_scan)
    ap.add_argument("--out", type=Path, default=default_out)
    ap.add_argument("--customer", default="Contoso Ltd.")
    args = ap.parse_args()

    # Load DMC scan (single-scan dir under root).
    scan_root = args.scan
    summary_files = list(scan_root.rglob("*.json"))
    top_summary_path = next(p for p in summary_files
                            if re.fullmatch(r"[0-9a-f-]{36}\.json", p.name)
                            and p.parent.parent == scan_root)
    scan = json.loads(top_summary_path.read_text())
    sub_id = scan["subscription_id"]
    sub_name = scan["subscription_name"] or "Contoso-Production-Hub"
    period_days = 90
    period_end = datetime.fromisoformat(scan["end_time"]).date().replace(day=1) - timedelta(days=1)
    period_start = period_end - timedelta(days=period_days - 1)
    currency = "USD"

    # Build per-VM rows (drive utilisation from the scan).
    rows = []
    for vm_summary in scan["vm_results"]:
        vm_uuid = vm_summary["vm_uuid"]
        mr_path = scan_root / scan["scan_id"] / sub_id / vm_uuid / "metric-result.json"
        if not mr_path.exists():
            continue
        payload = list(json.loads(mr_path.read_text()).values())[0]
        # Inject the metric_collection from inside the metric file.
        rows.extend(emit_vm_rows(payload, sub_id, sub_name, currency, period_days))

    rows.extend(emit_inventory_rows(scan, sub_id, sub_name, currency, period_days, "rg-contoso-prod-web-eus"))

    # Bias rng for noise: a -$1 rounding adjustment, mirroring real exports.
    rounding = round(rng.uniform(-2.5, -0.5), 6)
    rows.append(["", "", "", "", "", "", "RoundingAdjustment", "", rounding, rounding, currency])

    total = round(sum(r[9] for r in rows), 6)

    wb = openpyxl.Workbook()
    summary_ws = wb.active
    summary_ws.title = "Summary"
    write_summary(summary_ws, args.customer, sub_name, sub_id, period_start, period_end, total)
    data_ws = wb.create_sheet("Data")
    write_data(data_ws, rows)

    args.out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(args.out)

    by_service = {}
    for r in rows:
        if not r[5]:
            continue
        by_service[r[5]] = by_service.get(r[5], 0) + r[9]
    print(f"Customer:       {args.customer}")
    print(f"Subscription:   {sub_id}")
    print(f"Period:         {period_start} → {period_end} ({period_days} days)")
    print(f"Rows:           {len(rows):,}  (total {currency} {total:,.2f})")
    print(f"Output:         {args.out}")
    print()
    print("Top services:")
    for svc, c in sorted(by_service.items(), key=lambda x: -x[1])[:10]:
        print(f"  {svc:32s} {currency} {c:>14,.2f}")


if __name__ == "__main__":
    main()
